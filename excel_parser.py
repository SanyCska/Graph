from openpyxl import load_workbook


def parse(filename):
    """ Парсинг Excel-файлов.

    Args:
        filename: Имя файла для парсинга.

    Returns:
        dict: Словарь с ветками и узлами.
    """
    wb = load_workbook(filename, read_only=True, data_only=True)
    sheet = wb.get_active_sheet()

    found_nodes = set()

    # Берем область, в которой находятся нужные значения
    for cellObj in sheet['C3': sheet.max_row]:
        for cell in cellObj:
            value = cell.value
            if (value is not None) and (value not in found_nodes):
                found_nodes.add(value)

    branches = []
    for row_index in range(3, sheet.max_row + 1):
        node1 = sheet.cell(row=row_index, column=1).value
        node2 = sheet.cell(row=row_index, column=2).value
        if node1 in found_nodes and node2 in found_nodes:
            branches.append((node1, node2))

    return {
        'nodes': found_nodes,
        'branches': branches
    }
