#!/usr/bin/python
from os import chdir
from openpyxl import load_workbook


def load_file(filename):
    chdir('C:/Users/stupak_av/Projects/Graph/upload')
    wb = load_workbook(filename)
    sheet = wb.get_sheet_by_name('Лист1')

    # Создаем два списка для проверки выполнения условий:
    # 1) если есть повторяющиеся номера узлов в Excel-файле, то игнорировать их;
    # 2) если в ветке указан номер несуществующего узла, игнорировать данную ветку;
    check_nodes = list()
    check_edge = list()

    # Берем область, в которой находятся нужные значения
    for cellObj in sheet['C3': sheet.max_row]:
        for cell in cellObj:
            if cell.value not in check_nodes:
                # заполняем списки уникальными значениями узлов графа
                check_nodes.append(cell.value)
                check_edge.append(cell.value)

    edge_list = list()
    print(sheet.max_row)
    for each in range(3, sheet.max_row + 1):
        # Проверка условия 1
        if sheet.cell(row=each, column=3).value in check_nodes:
            # Проверка условия 2
            if sheet.cell(row=each, column=1).value in check_edge and sheet.cell(row=each,
                                                                                 column=2).value in check_edge:
                edge_list.append([sheet.cell(row=each, column=1).value, sheet.cell(row=each, column=2).value])
                check_nodes.remove(sheet.cell(row=each, column=3).value)

    return edge_list

# if __name__ == '__main__':
#     load_file('test1.xlsx')
